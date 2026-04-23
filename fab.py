#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from itertools import zip_longest
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_NAME = "Fab"
OUTPUT_FILE = Path.cwd() / "Fab.xlsx"

TARGET_OWNER = "jeunghun@isi.edu"
DEFAULT_SERVICE = "Fab Service"
DEFAULT_PIPELINE = "Fab [1. Customer Evaluation]"
DEFAULT_STAGE = "Export Control Check (Fab [1. Customer Evaluation])"

TARGET_HEADERS = [
    "Deal Name",
    "Company Name",
    "First name",
    "Last name",
    "Email",
    "Phone",
    "M2 Service",
    "Notes",
    "Company Domain Name",
    "Company Owner",
    "Deal Owner",
    "Contact Owner",
    "Pipeline",
    "Deal Stage",
]

HEADER_ALIASES: Dict[str, str] = {
    "company name": "company_name",
    "name": "name",
    "email": "email",
    "phone": "phone",
    "phone #": "phone",
    "phone number": "phone",
    "notes": "notes",
    "type": "type",
    "date": "date",
    "how?": "how",
    "how": "how",
    "update": "update",
    "description of needs": "description_of_needs",
    "description of need": "description_of_need",
    "intro email": "intro_email",
    "intro chat": "intro_chat",
    "nda cover memo": "nda_cover_memo",
    "nda": "nda",
    "service contract": "service_contract",
    "sow": "sow",
    "production started": "production_started",
    "production finished": "production_finished",
}


def find_input_workbook() -> Path:
    cwd = Path.cwd()
    candidates = [
        p for p in sorted(cwd.glob("*.xlsx"))
        if not p.name.startswith("~$") and p.name != OUTPUT_FILE.name
    ]
    if not candidates:
        raise FileNotFoundError("No input .xlsx file found in the current directory.")

    for path in candidates:
        try:
            wb = load_workbook(path, read_only=True)
            has_sheet = SHEET_NAME in wb.sheetnames
            wb.close()
            if has_sheet:
                return path
        except Exception:
            continue

    return candidates[0]


def raw_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", raw_text(value)).strip()


def norm(value: object) -> str:
    return clean_text(value).lower()


def is_blank(value: object) -> bool:
    return clean_text(value) == ""


def split_lines(value: object) -> List[str]:
    text = raw_text(value)
    if not text.strip():
        return []
    parts = re.split(r"\r\n|\r|\n", text)
    return [p.strip() for p in parts if p.strip()]


PREFIXES = {
    "mr", "mr.", "mrs", "mrs.", "ms", "ms.", "miss", "dr", "dr.",
    "prof", "prof.", "sir", "madam", "mx", "mx.", "rev", "rev.",
    "fr", "fr.", "hon", "hon.", "judge", "pres", "pres.", "gov", "gov.",
    "coach", "capt", "capt.", "captain",
}


def split_name(full_name: object) -> Tuple[str, str]:
    text = clean_text(full_name)
    if not text:
        return "", ""

    parts = text.split()
    if len(parts) == 1:
        return parts[0], ""

    first = parts[0]
    if first.lower().rstrip(".") in {p.rstrip(".") for p in PREFIXES} and len(parts) >= 2:
        return f"{first} {parts[1]}".strip(), " ".join(parts[2:]).strip()

    return parts[0], " ".join(parts[1:]).strip()


def infer_domain_from_company(company_name: object, email_value: object) -> str:
    """
    Infer a single domain name from the company name first.
    If that fails, use the domain part of the email address.
    Never return a full email address.
    """
    company = clean_text(company_name).lower()
    if company and "@" not in company:
        company = re.sub(
            r"\b(incorporated|inc|llc|ltd|limited|corp|corporation|co|company|plc|gmbh|sarl|pvt|pty)\b",
            "",
            company,
        )
        company = re.sub(r"[^a-z0-9]+", "", company)
        if company:
            return f"{company}.com"

    email = clean_text(email_value)
    if "@" in email:
        domain = email.split("@", 1)[1].strip().lower()
        domain = re.sub(r"^www\.", "", domain)
        return domain.strip(" .;,")
    return ""


def row_values(ws, row_idx: int, max_col: int) -> List[object]:
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


def row_has_strike_through(ws, row_idx: int) -> bool:
    cell = ws.cell(row=row_idx, column=1)
    try:
        return bool(cell.font and cell.font.strike)
    except Exception:
        return False


def is_checked(value: object) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return clean_text(value).lower() in {"true", "yes", "y", "x", "1", "checked", "on", "✓"}


class SourceColumns:
    def __init__(self) -> None:
        self.company_name: Optional[int] = None
        self.name: Optional[int] = None
        self.email: Optional[int] = None
        self.phone: Optional[int] = None
        self.notes: Optional[int] = None
        self.type: Optional[int] = None
        self.date: Optional[int] = None
        self.how: Optional[int] = None
        self.update: Optional[int] = None
        self.description_of_needs: Optional[int] = None
        self.description_of_need: Optional[int] = None

        self.intro_email: Optional[int] = None
        self.intro_chat: Optional[int] = None
        self.nda_cover_memo: Optional[int] = None
        self.nda: Optional[int] = None
        self.service_contract: Optional[int] = None
        self.sow: Optional[int] = None
        self.production_started: Optional[int] = None
        self.production_finished: Optional[int] = None


def build_source_columns(ws) -> SourceColumns:
    """
    Column label is found in row 2.
    If a cell in row 2 is empty, use row 1 to determine the label.
    Only rows 1 and 2 are used for label detection.
    """
    cols = SourceColumns()
    for col_idx in range(1, ws.max_column + 1):
        top = clean_text(ws.cell(1, col_idx).value)
        second = clean_text(ws.cell(2, col_idx).value)
        header = second or top
        key = norm(header)
        if key in HEADER_ALIASES:
            setattr(cols, HEADER_ALIASES[key], col_idx)
    return cols


def compose_notes(ws_values, row_idx: int, cols: SourceColumns) -> str:
    parts: List[str] = []

    def add_text(label: str, col_idx: Optional[int]) -> None:
        if col_idx is None:
            return
        v = ws_values.cell(row=row_idx, column=col_idx).value
        if not is_blank(v):
            parts.append(f"{label}: {clean_text(v)}")

    add_text("Description of Needs", cols.description_of_needs)
    add_text("Description of Need", cols.description_of_need)
    add_text("Type", cols.type)

    if cols.date is not None:
        v = ws_values.cell(row=row_idx, column=cols.date).value
        if not is_blank(v):
            parts.append(f"First contact Date: {clean_text(v)}")

    if cols.how is not None:
        v = ws_values.cell(row=row_idx, column=cols.how).value
        if not is_blank(v):
            parts.append(f"First contact How: {clean_text(v)}")

    if cols.update is not None:
        v = ws_values.cell(row=row_idx, column=cols.update).value
        if not is_blank(v):
            parts.append(f"Update: {clean_text(v)}")

    tracker_fields = [
        ("Intro Email", cols.intro_email),
        ("Intro Chat", cols.intro_chat),
        ("NDA Cover Memo", cols.nda_cover_memo),
        ("NDA", cols.nda),
        ("Service Contract", cols.service_contract),
        ("SOW", cols.sow),
        ("Production Started", cols.production_started),
        ("Production Finished", cols.production_finished),
    ]

    checked_items: List[str] = []
    for label, col_idx in tracker_fields:
        if col_idx is None:
            continue
        if is_checked(ws_values.cell(row=row_idx, column=col_idx).value):
            checked_items.append(label)

    if checked_items:
        parts.append("Tracker: " + ", ".join(checked_items))

    if cols.production_finished is not None:
        extras: List[str] = []
        for c in range(cols.production_finished + 1, ws_values.max_column + 1):
            v = ws_values.cell(row=row_idx, column=c).value
            if not is_blank(v):
                extras.append(clean_text(v))
        if extras:
            parts.append(" | ".join(extras))

    return " | ".join(parts)


def split_contacts(name_value: object, email_value: object) -> Tuple[List[Tuple[str, str, str]], int, int]:
    names = split_lines(name_value)
    emails = split_lines(email_value)

    multi_name_cells = 1 if len(names) >= 2 else 0
    multi_email_cells = 1 if len(emails) >= 2 else 0

    if not names:
        names = [""]
    if not emails:
        emails = [""]

    contacts: List[Tuple[str, str, str]] = []
    for name_text, email_text in zip_longest(names, emails, fillvalue=None):
        if name_text is None:
            name_text = names[-1]
        if email_text is None:
            email_text = emails[-1]
        first_name, last_name = split_name(name_text)
        contacts.append((first_name, last_name, clean_text(email_text)))

    return contacts, multi_name_cells, multi_email_cells


def style_output_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(TARGET_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(header)) + 2, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TARGET_HEADERS))}1"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(TARGET_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def process_sheet(input_file: Path, sheet_name: str, output_file: Path) -> None:
    wb_style = load_workbook(input_file, data_only=False)
    wb_values = load_workbook(input_file, data_only=True)

    ws_style = wb_style[sheet_name]
    ws_values = wb_values[sheet_name]

    max_row = ws_style.max_row
    max_col = ws_style.max_column

    source_cols = build_source_columns(ws_values)
    if source_cols.company_name is None:
        raise ValueError("Could not find the 'Company Name' column.")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name

    ignored_rows = 0
    copied_rows = 0
    multi_name_cells = 0
    multi_email_cells = 0
    out_rows: List[List[object]] = []

    label_row_1 = [norm(v) for v in row_values(ws_values, 1, max_col)]
    label_row_2 = [norm(v) for v in row_values(ws_values, 2, max_col)]

    for row_idx in range(3, max_row + 1):
        row_vals = row_values(ws_values, row_idx, max_col)
        if all(is_blank(v) for v in row_vals):
            continue

        first_cell = ws_values.cell(row=row_idx, column=1).value
        if is_blank(first_cell):
            ignored_rows += 1
            continue

        if row_has_strike_through(ws_style, row_idx):
            ignored_rows += 1
            continue

        first_text = clean_text(first_cell).lower()
        if first_text == "company name":
            ignored_rows += 1
            continue

        current = [norm(v) for v in row_vals]
        if current == label_row_1 or current == label_row_2:
            ignored_rows += 1
            continue

        company_name = clean_text(ws_values.cell(row=row_idx, column=source_cols.company_name).value)
        if not company_name:
            ignored_rows += 1
            continue

        print(f"Company Name: {company_name}")

        name_value = ws_values.cell(row=row_idx, column=source_cols.name).value if source_cols.name else ""
        email_value = ws_values.cell(row=row_idx, column=source_cols.email).value if source_cols.email else ""

        contacts, name_multi_flag, email_multi_flag = split_contacts(name_value, email_value)
        multi_name_cells += name_multi_flag
        multi_email_cells += email_multi_flag

        phone = clean_text(ws_values.cell(row=row_idx, column=source_cols.phone).value if source_cols.phone else "")
        notes = compose_notes(ws_values, row_idx, source_cols)
        company_domain = infer_domain_from_company(company_name, email_value)

        for first_name, last_name, email in contacts:
            out_rows.append([
                company_name,          # Deal Name
                company_name,          # Company Name
                first_name,
                last_name,
                email,
                phone,
                DEFAULT_SERVICE,
                notes,
                company_domain,
                TARGET_OWNER,          # Company Owner
                TARGET_OWNER,          # Deal Owner
                TARGET_OWNER,          # Contact Owner
                DEFAULT_PIPELINE,
                DEFAULT_STAGE,
            ])
            copied_rows += 1

    for c_idx, header in enumerate(TARGET_HEADERS, start=1):
        out_ws.cell(row=1, column=c_idx, value=header)

    for r_idx, row in enumerate(out_rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            out_ws.cell(row=r_idx, column=c_idx, value=value)

    style_output_sheet(out_ws)
    out_wb.save(output_file)

    print(f"{sheet_name}: ignored {ignored_rows} rows, copied {copied_rows} rows -> {output_file.name}")
    print(f"Rows with multiple names: {multi_name_cells}")
    print(f"Rows with multiple emails: {multi_email_cells}")


def main() -> int:
    try:
        input_file = find_input_workbook()
    except FileNotFoundError as e:
        print(str(e), file=sys.stderr)
        return 1

    wb = load_workbook(input_file, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet not found: {SHEET_NAME}", file=sys.stderr)
        return 1

    process_sheet(input_file, SHEET_NAME, OUTPUT_FILE)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
