from __future__ import annotations

from pathlib import Path
import re
from datetime import datetime, date
from itertools import zip_longest

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment


INPUT_FILE = Path("./External Customers - Engagement Tracking.xlsx")
OUTPUT_FILE = Path("./IMS2025.xlsx")
SOURCE_SHEET_NAME = "IMS 2025"
OWNER_EMAIL = "jeunghun@isi.edu"

# The workbook column is named "Create Date".
# "Date Added" will be copied there when it is a real date.
TARGET_HEADERS = [
    "Deal Name",
    "Pipeline",
    "Deal Stage",
    "Create Date",
    "Company Name",
    "First name",
    "Last name",
    "Email",
    "Phone",
    "Position",
    "Notes",
    "Company Domain Name",
    "Company Owner",
    "Contact Owner",
]

INTEREST_COLUMNS = ["Fab Services", "MPW", "IC Design", "EDA Tool Training", "Other"]
WHO_MET_COLUMNS = ["Jeung", "Christine", "Xi"]


def clean(value):
    return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", clean(value)).lower().strip()


def header_label(ws, col_idx):
    """
    Column label is found in row 2.
    If row 2 is empty, use row 1.
    """
    row2 = clean(ws.cell(2, col_idx).value)
    row1 = clean(ws.cell(1, col_idx).value)
    return row2 if row2 else row1


def find_col(ws, wanted_labels):
    wanted_norms = [norm(x) for x in wanted_labels]
    for c in range(1, ws.max_column + 1):
        label = header_label(ws, c)
        n = norm(label)
        if not n:
            continue
        for wn in wanted_norms:
            if n == wn or n.startswith(wn):
                return c
    return None


def has_strikethrough(cell):
    return bool(cell.font and cell.font.strike)


def is_date_like(cell):
    if cell is None:
        return False
    if isinstance(cell.value, (datetime, date)):
        return True
    return bool(getattr(cell, "is_date", False))


def is_checked(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = norm(value)
    return s in {"1", "true", "yes", "y", "x", "checked", "check", "✓", "☑", "on"}


def split_multiline(value):
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"\r\n|\r|\n", text) if part.strip()]


def split_name(name_text):
    """
    Keep prefixes such as Prof, Mr., Dr. with the first name.
    Examples:
        "Prof John Smith" -> ("Prof John", "Smith")
        "Smith, John"     -> ("John", "Smith")
    """
    text = clean(name_text)
    if not text:
        return "", ""

    if "," in text:
        last_name, first_name = [p.strip() for p in text.split(",", 1)]
        return first_name, last_name

    parts = text.split()
    if len(parts) == 1:
        return parts[0], ""

    return " ".join(parts[:-1]), parts[-1]


def infer_company_domain(company_name, email):
    """
    Infer company domain from company name first, then email domain.
    Never return an email address as the domain.
    """
    company_text = clean(company_name)
    email_text = clean(email)

    # Company-name-based inference first.
    if company_text and "@" not in company_text:
        candidate = company_text.lower().strip()
        candidate = candidate.replace("https://", "").replace("http://", "")
        candidate = candidate.replace("www.", "")
        candidate = candidate.strip("/ ")

        # If it already looks like a domain, use it.
        if "." in candidate and " " not in candidate:
            return candidate

        slug = re.sub(r"[^a-z0-9]+", "", candidate)
        if slug:
            return f"{slug}.com"

    # Fallback: email domain only.
    m = re.search(r"@([A-Za-z0-9.-]+\.[A-Za-z]{2,})", email_text)
    if m:
        domain = m.group(1).lower().strip(".")
        domain = re.sub(r"^(mail|www|mx|smtp|webmail)\.", "", domain)
        if domain:
            return domain

    return ""


def style_header(ws):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True)
    for i, header in enumerate(TARGET_HEADERS, start=1):
        cell = ws.cell(1, i)
        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def row_is_blank(ws, row_idx):
    return all(
        ws.cell(row_idx, c).value is None or str(ws.cell(row_idx, c).value).strip() == ""
        for c in range(1, ws.max_column + 1)
    )


def get_source_rows(ws):
    # Data begins after the two header rows.
    for r in range(3, ws.max_row + 1):
        yield r


def convert():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    wb = load_workbook(INPUT_FILE)
    if SOURCE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SOURCE_SHEET_NAME}' was not found.")

    ws = wb[SOURCE_SHEET_NAME]

    # Source columns
    date_added_col = find_col(ws, ["Date Added"])
    org_col = find_col(ws, ["Organization"])
    name_col = find_col(ws, ["Name"])
    first_name_col = find_col(ws, ["First Name", "Firstname"])
    last_name_col = find_col(ws, ["Last Name", "Lastname"])
    email_col = find_col(ws, ["Email", "E-mail"])
    phone_col = find_col(ws, ["Phone #", "Phone"])
    comments_col = find_col(ws, ["Comments"])
    assigned_to_col = find_col(ws, ["Assigned To"])
    contacted_col = find_col(ws, ["Contacted?"])
    next_steps_col = find_col(ws, ["Next Steps"])

    position_col = None
    for c in range(1, ws.max_column + 1):
        label = header_label(ws, c)
        if norm(label).startswith("verified"):
            position_col = c
            break

    interest_cols = {label: find_col(ws, [label]) for label in INTEREST_COLUMNS}
    who_met_cols = {label: find_col(ws, [label]) for label in WHO_MET_COLUMNS}

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = SOURCE_SHEET_NAME
    style_header(out_ws)
    out_ws.freeze_panes = "A2"

    ignored_rows = 0
    copied_rows = 0
    multi_name_cells = 0
    multi_email_cells = 0

    out_row = 2

    for r in get_source_rows(ws):
        # Ignore rows if first cell is empty
        first_cell = ws.cell(r, 1)
        if first_cell.value is None or str(first_cell.value).strip() == "":
            ignored_rows += 1
            continue

        # Ignore rows whose first cell has strike-through
        if has_strikethrough(first_cell):
            ignored_rows += 1
            continue

        # Ignore fully blank rows
        if row_is_blank(ws, r):
            ignored_rows += 1
            continue

        company_name = clean(ws.cell(r, org_col).value) if org_col else ""
        if not company_name:
            ignored_rows += 1
            continue

        # Create Date comes from Date Added if it is a real date.
        # If Date Added is not a date but has content, append it to Notes.
        create_date = ""
        date_added_note = ""
        if date_added_col:
            c = ws.cell(r, date_added_col)
            if is_date_like(c):
                create_date = c.value
            else:
                date_added_text = clean(c.value)
                if date_added_text:
                    date_added_note = f"Date Added: {date_added_text}"

        # Position from a column whose label starts with "Verified"
        position = clean(ws.cell(r, position_col).value) if position_col else ""

        # Name handling
        if first_name_col and last_name_col:
            first_vals = split_multiline(ws.cell(r, first_name_col).value)
            last_vals = split_multiline(ws.cell(r, last_name_col).value)

            if len(first_vals) > 1 or len(last_vals) > 1:
                multi_name_cells += 1

            if not first_vals:
                first_vals = [""]
            if not last_vals:
                last_vals = [""]

            name_pairs = list(zip_longest(first_vals, last_vals, fillvalue=""))
            name_pairs = [(clean(fn), clean(ln)) for fn, ln in name_pairs]
        else:
            raw_names = split_multiline(ws.cell(r, name_col).value if name_col else "")
            if len(raw_names) > 1:
                multi_name_cells += 1
            if not raw_names:
                raw_names = [""]
            name_pairs = [split_name(x) for x in raw_names]

        # Email handling
        email_values = split_multiline(ws.cell(r, email_col).value if email_col else "")
        if len(email_values) > 1:
            multi_email_cells += 1
        if not email_values:
            email_values = [""]

        max_len = max(len(name_pairs), len(email_values))
        if len(name_pairs) == 1 and max_len > 1:
            name_pairs = name_pairs * max_len
        if len(email_values) == 1 and max_len > 1:
            email_values = email_values * max_len

        paired_rows = []
        for i in range(max_len):
            fn, ln = name_pairs[i] if i < len(name_pairs) else ("", "")
            em = email_values[i] if i < len(email_values) else ""
            paired_rows.append((fn, ln, em))

        phone = clean(ws.cell(r, phone_col).value) if phone_col else ""

        # Company domain from company name first, then email domain.
        domain_seed_email = email_values[0] if email_values else ""
        company_domain = infer_company_domain(company_name, domain_seed_email)

        # Notes assembly
        notes_parts = []

        # Interested in
        interested_labels = [
            label for label, col in interest_cols.items()
            if col and is_checked(ws.cell(r, col).value)
        ]
        if interested_labels:
            notes_parts.append("Interested in: " + ", ".join(interested_labels))

        # Comments
        comments = clean(ws.cell(r, comments_col).value) if comments_col else ""
        if comments:
            notes_parts.append(f"Comments: {comments}")

        # Who met?
        who_met_labels = [
            label for label, col in who_met_cols.items()
            if col and is_checked(ws.cell(r, col).value)
        ]
        if who_met_labels:
            notes_parts.append("Who met?: " + ", ".join(who_met_labels))

        # Assigned To
        assigned_to = clean(ws.cell(r, assigned_to_col).value) if assigned_to_col else ""
        if assigned_to:
            notes_parts.append(f"Follow-up: Assigned to {assigned_to}")
        else:
            notes_parts.append("Follow-up: Assigned to NONE")

        # Contacted?
        if contacted_col:
            contacted_value = "Yes" if is_checked(ws.cell(r, contacted_col).value) else "No"
            notes_parts.append(f"Contacted?: {contacted_value}")

        # Next Steps
        next_steps = clean(ws.cell(r, next_steps_col).value) if next_steps_col else ""
        if next_steps:
            notes_parts.append(f"Next Steps: {next_steps}")

        # Date Added text when not a date
        if date_added_note:
            notes_parts.append(date_added_note)

        notes = " | ".join(notes_parts)

        company_owner = OWNER_EMAIL
        contact_owner = OWNER_EMAIL

        for first_name, last_name, email in paired_rows:
            # Deal Name = Company Name + "-" + First Name + "-" + Last Name + "-IMS 2025"
            deal_name = f"{company_name}-{first_name}-{last_name}-IMS 2025"

            out_ws.cell(out_row, 1).value = deal_name
            out_ws.cell(out_row, 2).value = "M2 Initial Contact"
            out_ws.cell(out_row, 3).value = "Initial Contact (M2 Initial Contact)"
            out_ws.cell(out_row, 4).value = create_date
            out_ws.cell(out_row, 5).value = company_name
            out_ws.cell(out_row, 6).value = first_name
            out_ws.cell(out_row, 7).value = last_name
            out_ws.cell(out_row, 8).value = email
            out_ws.cell(out_row, 9).value = phone
            out_ws.cell(out_row, 10).value = position
            out_ws.cell(out_row, 11).value = notes
            out_ws.cell(out_row, 12).value = company_domain
            out_ws.cell(out_row, 13).value = company_owner
            out_ws.cell(out_row, 14).value = contact_owner

            out_row += 1
            copied_rows += 1

    out_ws.auto_filter.ref = f"A1:N{max(out_row - 1, 1)}"
    autosize_columns(out_ws)
    out_wb.save(OUTPUT_FILE)

    print(f"Saved: {OUTPUT_FILE}")
    print(f"Ignored rows: {ignored_rows}")
    print(f"Copied rows: {copied_rows}")
    print(f"Cells with multiple names: {multi_name_cells}")
    print(f"Cells with multiple emails: {multi_email_cells}")


if __name__ == "__main__":
    convert()
